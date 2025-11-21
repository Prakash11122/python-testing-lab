import openpyxl

#File-->Workbook--->Sheets-->Rows-->Cells

file="/Users/prakashpradhan/Downloads/Python_Testing/Selenium/ Data_Driven_Testing_Excel_class_14/ClassExamples/caldata.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row # count number of rows in a excel sheet
cols=sheet.max_column #count number of columns in a excel sheet

#Reading all the rows & columns from excel sheet
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='                ')
    print()







